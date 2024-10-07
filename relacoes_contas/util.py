import openpyxl
import json

def criar_plano_contas():
    wb = openpyxl.load_workbook('/Users/matheusbarbosa/Downloads/plano_contas.xlsx')
    ws = wb.active

    plano_contas = []

    for row in range(2, ws.max_row + 1):
        cod_cta = ws[f"A{row}"].value
        clas_cta = ws[f"B{row}"].value
        tipo_cta = ws[f"C{row}"].value
        nome_cta = ws[f"D{row}"].value
        conta = {}
        conta['id'] = cod_cta
        conta['classificacao'] = clas_cta
        conta['tipo'] = tipo_cta
        conta['descricao'] = nome_cta

        plano_contas.append(conta)
        print(cod_cta, clas_cta, tipo_cta, nome_cta)

    with open('plano_contas.json', 'w', encoding='utf-8') as arquivo:
        json.dump(plano_contas, arquivo, ensure_ascii=False, indent=4)

def plano_contas():
    with open('relacoes_contas/plano_contas.json', 'r', encoding='utf-8') as arquivo:
        dados_carregados = json.load(arquivo)
    return dados_carregados

def bancos():
    with open('relacoes_contas/bancos.json', 'r', encoding='utf-8') as arquivo:
        dados_carregados = json.load(arquivo)
    return dados_carregados


plano_contas()